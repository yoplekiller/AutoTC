"""
Jira 티켓 기반 매뉴얼 TC 자동 생성

사용법:
  # 단일 티켓 (티켓 키 또는 URL)
  python src/generate_tc.py PROJ-123
  python src/generate_tc.py https://yourcompany.atlassian.net/browse/PROJ-123

  # 서비스 컨텍스트 적용
  python src/generate_tc.py PROJ-123 --context kream

  # 엑셀 일괄 처리 (A열에 티켓 URL/키 목록)
  python src/generate_tc.py tickets.xlsx --context kream

  # 구글 스프레드시트 일괄 처리
  python src/generate_tc.py https://docs.google.com/spreadsheets/d/SHEET_ID/edit

  # 입력용 템플릿 엑셀 생성
  python src/generate_tc.py --template
"""

import sys
import io
import re
import os
import json
import argparse
from datetime import datetime

if sys.stdout.encoding is None or sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from jira import JIRA
from groq import Groq
from dotenv import load_dotenv
from tc_core import (
    DailyTokenLimitError,
    extract_issue_key, fetch_issue, load_context,
    augment_ticket_spec, analyze_spec_for_plan, generate_test_cases,
    filter_tc_list, dedupe_tc_list, _get_gspread_client,
)

load_dotenv()

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CREDS_PATH = os.path.join(ROOT_DIR, os.getenv("GOOGLE_CREDENTIALS_PATH", "credentials.json"))
JIRA_URL = os.getenv("JIRA_URL", "")


# ── 엑셀 입력/출력 ────────────────────────────────────────────────────

def create_template(output_path: str = "tickets_template.xlsx"):
    """입력용 티켓 목록 템플릿 엑셀을 생성합니다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "티켓목록"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="티켓 URL 또는 이슈 키").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = header_align
    ws.cell(row=1, column=2, value="메모 (선택)").font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    ws.cell(row=1, column=2).alignment = header_align
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 30
    ws.row_dimensions[1].height = 22

    examples = [
        ("MKQA-1", "로그인 기능 TC"),
        ("MKQA-2", "회원가입 TC"),
        (f"{JIRA_URL}/browse/MKQA-3", "URL 형식 예시"),
    ]
    for r, (key, memo) in enumerate(examples, start=2):
        ws.cell(row=r, column=1, value=key).font = Font(color="808080", italic=True)
        ws.cell(row=r, column=2, value=memo).font = Font(color="808080", italic=True)

    wb.save(output_path)
    return output_path


def read_keys_from_excel(file_path: str) -> list:
    """엑셀 A열에서 티켓 URL/키 목록을 읽어옵니다. (헤더 제외, 빈 셀 스킵)"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    keys = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if val and str(val).strip():
            keys.append(str(val).strip())
    return keys


def save_excel(results: list, output_path: str):
    """결과를 엑셀 파일로 저장합니다. 티켓마다 별도 시트 생성."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 빈 시트 제거

    headers = [
        "TC ID", "대분류", "소분류", "테스트 유형",
        "테스트 시나리오(목적)", "사전 조건", "테스트 단계", "기대 결과",
        "테스트 상태", "비고 / 버그 링크", "우선순위", "위험도", "자동화 가능여부",
        "Quality 판정", "판정 사유",
        "요구사항 근거", "요구사항 상태", "실행 가능성", "필요 도구", "실행 메모", "설계 기법",
    ]
    col_widths = [14, 14, 16, 12, 35, 28, 45, 35, 12, 20, 10, 10, 14, 12, 32, 16, 16, 20, 20, 30, 16]
    last_col_letter = "U"
    # Quality 판정(N)/판정 사유(O)는 QA가 후보 TC를 훑어볼 때 바로 봐야 하는 핵심 신호라 노출.
    # 나머지(요구사항 근거~설계 기법)는 판정 근거를 따져볼 때만 필요한 메타데이터라 M열과 같은 방식으로 숨김.
    hidden_columns = ["M", "P", "Q", "R", "S", "T", "U"]

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note_align   = Alignment(horizontal="left", vertical="top", wrap_text=True)

    level_fills = {
        "High":   PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
        "Medium": PatternFill(start_color="FFF3CC", end_color="FFF3CC", fill_type="solid"),
        "Low":    PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
    }
    quality_fills = {
        "PASS":   PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        "REVIEW": PatternFill(start_color="FFF3CC", end_color="FFF3CC", fill_type="solid"),
        "REJECT": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    }

    for item in results:
        sheet_title = re.sub(r'[\\/*?:\[\]]', '', item["summary"])[:31]
        ws = wb.create_sheet(title=sheet_title)

        # 행 1: 티켓 URL 정보 (하이퍼링크)
        # item에 "url"이 명시되어 있으면(예: 기획서 기반 생성의 Confluence 페이지) 그걸 우선 쓰고,
        # 없으면 기존처럼 Jira 티켓 링크로 조립한다. 링크가 없는 경우(로컬 기획서 파일 등)는 하이퍼링크를 걸지 않는다.
        ticket_url = item.get("url", f"{JIRA_URL}/browse/{item['key']}")
        info_cell = ws.cell(row=1, column=1, value=f"{item['key']}  |  {item['summary']}")
        if ticket_url:
            info_cell.hyperlink = ticket_url
        info_cell.font = Font(bold=True, color="0563C1", underline="single", size=11)
        info_cell.fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
        info_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"A1:{last_col_letter}1")
        ws.row_dimensions[1].height = 22

        # 행 2: 요구사항 분석 (확인 필요한 질문 포함)
        note_cell = ws.cell(row=2, column=1, value=item.get("augmented_spec", ""))
        note_cell.font = Font(size=10, color="444444")
        note_cell.alignment = note_align
        ws.merge_cells(f"A2:{last_col_letter}2")
        ws.row_dimensions[2].height = 140

        # 행 3: 컬럼 헤더
        for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[3].height = 25
        # 자동화가능여부(M) + 요구사항/실행성/설계기법 메타데이터(P~U)는 수동 테스트 수행 시
        # 방해되지 않도록 기본 숨김 처리 (필요할 때만 펼쳐서 확인)
        for col_letter in hidden_columns:
            ws.column_dimensions[col_letter].hidden = True

        # 행 4~: TC 데이터
        last_row = 3 + len(item["test_cases"])
        for r_idx, tc in enumerate(item["test_cases"], start=4):
            ws.cell(row=r_idx, column=1,  value=tc.get("tc_id", "")).alignment = data_align
            ws.cell(row=r_idx, column=2,  value=tc.get("대분류", "")).alignment = data_align
            ws.cell(row=r_idx, column=3,  value=tc.get("소분류", "")).alignment = data_align
            ws.cell(row=r_idx, column=4,  value=tc.get("테스트유형", "")).alignment = data_align
            ws.cell(row=r_idx, column=5,  value=tc.get("테스트시나리오", "")).alignment = data_align
            ws.cell(row=r_idx, column=6,  value=tc.get("사전조건", "")).alignment = data_align
            ws.cell(row=r_idx, column=7,  value=tc.get("테스트단계", "")).alignment = data_align
            ws.cell(row=r_idx, column=8,  value=tc.get("기대결과", "")).alignment = data_align
            ws.cell(row=r_idx, column=9,  value="").alignment = data_align   # 테스트 상태
            ws.cell(row=r_idx, column=10, value="").alignment = data_align   # 연결 버그/비고
            priority = tc.get("우선순위", "")
            p_cell = ws.cell(row=r_idx, column=11, value=priority)
            p_cell.alignment = data_align
            if priority in level_fills:
                p_cell.fill = level_fills[priority]
            risk = tc.get("위험도", "")
            r_cell = ws.cell(row=r_idx, column=12, value=risk)
            r_cell.alignment = data_align
            if risk in level_fills:
                r_cell.fill = level_fills[risk]
            ws.cell(row=r_idx, column=13, value=tc.get("자동화가능여부", "")).alignment = data_align
            quality_status = tc.get("quality_status", "")
            q_cell = ws.cell(row=r_idx, column=14, value=quality_status)
            q_cell.alignment = data_align
            if quality_status in quality_fills:
                q_cell.fill = quality_fills[quality_status]
            ws.cell(row=r_idx, column=15, value=tc.get("quality_reason", "")).alignment = data_align
            ws.cell(row=r_idx, column=16, value=", ".join(tc.get("requirement_refs", []))).alignment = data_align
            ws.cell(row=r_idx, column=17, value=tc.get("requirement_status", "")).alignment = data_align
            ws.cell(row=r_idx, column=18, value=tc.get("execution_type", "")).alignment = data_align
            ws.cell(row=r_idx, column=19, value=", ".join(tc.get("required_tools", []))).alignment = data_align
            ws.cell(row=r_idx, column=20, value=tc.get("execution_note", "")).alignment = data_align
            ws.cell(row=r_idx, column=21, value=tc.get("test_design_technique", "")).alignment = data_align
            ws.row_dimensions[r_idx].height = 70

        # TC가 0개(예: 한도 초과로 생성 중단)면 드롭다운을 걸 데이터 범위 자체가 없으므로 스킵
        if item["test_cases"]:
            # 테스트 상태 드롭다운 (I열): P / F / B / N/A
            dv_status = DataValidation(type="list", formula1='"P,F,B,N/A"', allow_blank=True, showDropDown=False)
            dv_status.sqref = f"I4:I{last_row}"
            ws.add_data_validation(dv_status)

            # 자동화 가능여부 드롭다운 (M열): 가능 / 불가능
            dv_auto = DataValidation(type="list", formula1='"가능,불가능"', allow_blank=True, showDropDown=False)
            dv_auto.sqref = f"M4:M{last_row}"
            ws.add_data_validation(dv_auto)

            # Quality 판정 드롭다운 (N열): QA가 AI 판정을 직접 덮어쓸 수 있도록 함
            dv_quality = DataValidation(type="list", formula1='"PASS,REVIEW,REJECT"', allow_blank=True, showDropDown=False)
            dv_quality.sqref = f"N4:N{last_row}"
            ws.add_data_validation(dv_quality)

    wb.save(output_path)


# ── 구글 스프레드시트 입력/출력 ──────────────────────────────────────

def extract_sheet_id(url: str) -> str:
    """구글 스프레드시트 URL에서 spreadsheet ID를 추출합니다."""
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if match:
        return match.group(1)
    # ID 직접 입력된 경우 (URL이 아닌 순수 ID)
    if re.fullmatch(r"[a-zA-Z0-9_-]{30,}", url):
        return url
    raise ValueError(f"유효한 구글 스프레드시트 URL이 아닙니다: {url}")


def extract_gid(url: str) -> str | None:
    """구글 스프레드시트 URL에서 특정 탭의 gid를 추출합니다. 없으면 None."""
    match = re.search(r"[#&]gid=(\d+)", url)
    return match.group(1) if match else None


def read_keys_from_sheets(sheet_id: str, gid: str | None = None) -> list:
    """구글 스프레드시트 A열에서 티켓 URL/키를 읽습니다. (헤더 제외)

    URL에 gid가 명시되어 있으면 그 탭을, 없으면 첫 번째 탭을 읽습니다.
    """
    gc = _get_gspread_client()
    sh = gc.open_by_key(sheet_id)
    if gid is not None:
        ws = sh.get_worksheet_by_id(int(gid))
    else:
        ws = sh.get_worksheet(0)
    all_values = ws.col_values(1)  # A열 전체
    # 헤더(1행) 제외, 빈 값 스킵
    return [v.strip() for v in all_values[1:] if v and v.strip()]


def save_to_sheets(results: list, sheet_id: str):
    """생성된 TC를 구글 스프레드시트에 티켓별 시트로 저장합니다."""
    import gspread

    gc = _get_gspread_client()
    sh = gc.open_by_key(sheet_id)

    headers = [
        "TC ID", "대분류", "소분류", "테스트 유형",
        "테스트 시나리오(목적)", "사전 조건", "테스트 단계", "기대 결과",
        "테스트 상태", "비고 / 버그 링크", "우선순위", "위험도", "자동화 가능여부",
        "Quality 판정", "판정 사유",
        "요구사항 근거", "요구사항 상태", "실행 가능성", "필요 도구", "실행 메모", "설계 기법",
    ]
    level_colors = {
        "High":   {"red": 1.0,  "green": 0.8,  "blue": 0.8},
        "Medium": {"red": 1.0,  "green": 0.95, "blue": 0.8},
        "Low":    {"red": 0.85, "green": 0.92, "blue": 0.85},
    }
    quality_colors = {
        "PASS":   {"red": 0.85, "green": 0.92, "blue": 0.85},
        "REVIEW": {"red": 1.0,  "green": 0.95, "blue": 0.8},
        "REJECT": {"red": 1.0,  "green": 0.8,  "blue": 0.8},
    }
    col_widths = [100, 110, 120, 110, 280, 200, 320, 260, 100, 160, 90, 90, 110, 90, 240, 130, 120, 140, 160, 220, 120]
    last_col_letter = "U"
    total_tc = 0

    for item in results:
        sheet_title = item["summary"][:100]
        # save_excel과 동일한 규칙: item["url"]이 있으면(기획서/Confluence 기반) 그걸 쓰고, 없으면 Jira 링크로 조립
        ticket_url = item.get("url", f"{JIRA_URL}/browse/{item['key']}")

        try:
            ws = sh.worksheet(sheet_title)
            ws.clear()
            print(f"  '{sheet_title}' 시트 초기화")
        except gspread.exceptions.WorksheetNotFound:
            ws = sh.add_worksheet(title=sheet_title, rows=200, cols=len(headers))
            print(f"  '{sheet_title}' 시트 생성")

        # 행 1: 티켓 URL 정보
        header_text = f"{item['key']}  |  {item['summary']}  |  {ticket_url}" if ticket_url else f"{item['key']}  |  {item['summary']}"
        ws.update([[header_text]], "A1")
        ws.format("A1", {
            "backgroundColor": {"red": 0.922, "green": 0.953, "blue": 0.984},
            "textFormat": {"bold": True, "foregroundColor": {"red": 0.02, "green": 0.34, "blue": 0.71}},
            "horizontalAlignment": "LEFT",
        })
        ws.merge_cells(f"A1:{last_col_letter}1")

        # 행 2: 요구사항 분석 (확인 필요한 질문 포함)
        ws.update([[item.get("augmented_spec", "")]], "A2")
        ws.format("A2", {
            "verticalAlignment": "TOP",
            "horizontalAlignment": "LEFT",
            "wrapStrategy": "WRAP",
            "textFormat": {"fontSize": 9, "foregroundColor": {"red": 0.27, "green": 0.27, "blue": 0.27}},
        })
        ws.merge_cells(f"A2:{last_col_letter}2")

        # 행 3: 컬럼 헤더
        ws.update([headers], "A3")
        ws.format(f"A3:{last_col_letter}3", {
            "backgroundColor": {"red": 0.267, "green": 0.447, "blue": 0.769},
            "textFormat": {"bold": True, "foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0}},
            "horizontalAlignment": "CENTER",
        })

        # 행 4~: TC 데이터
        rows_data = []
        for tc in item["test_cases"]:
            rows_data.append([
                tc.get("tc_id", ""),
                tc.get("대분류", ""),
                tc.get("소분류", ""),
                tc.get("테스트유형", ""),
                tc.get("테스트시나리오", ""),
                tc.get("사전조건", ""),
                tc.get("테스트단계", ""),
                tc.get("기대결과", ""),
                "",  # 테스트 상태
                "",  # 연결 버그/비고
                tc.get("우선순위", ""),
                tc.get("위험도", ""),
                tc.get("자동화가능여부", ""),
                tc.get("quality_status", ""),
                tc.get("quality_reason", ""),
                ", ".join(tc.get("requirement_refs", [])),
                tc.get("requirement_status", ""),
                tc.get("execution_type", ""),
                ", ".join(tc.get("required_tools", [])),
                tc.get("execution_note", ""),
                tc.get("test_design_technique", ""),
            ])
        if rows_data:
            ws.update(rows_data, "A4")
            end_row = 4 + len(rows_data)

            # 데이터 셀 정렬: 세로=가운데, 가로=왼쪽
            ws.format(f"A4:{last_col_letter}{end_row - 1}", {
                "verticalAlignment": "MIDDLE",
                "horizontalAlignment": "LEFT",
                "wrapStrategy": "WRAP",
            })

            # 우선순위(K열)/위험도(L열) 색상 — 셀 하나당 ws.format() 호출 1번씩 하면 TC가 많을 때
            # (기획서 기반 생성처럼 40~50개 이상 나오는 경우) 분당 쓰기 요청 할당량(429)을 바로 초과한다.
            # 실제로 --sheet 옵션 실전 검증 중 57개 TC로 재현됨. batch_format으로 모아서 한 번에 보낸다.
            color_requests = []
            for i, tc in enumerate(item["test_cases"]):
                p_color = level_colors.get(tc.get("우선순위", ""))
                if p_color:
                    color_requests.append({"range": f"K{4 + i}", "format": {"backgroundColor": p_color}})
                r_color = level_colors.get(tc.get("위험도", ""))
                if r_color:
                    color_requests.append({"range": f"L{4 + i}", "format": {"backgroundColor": r_color}})
                q_color = quality_colors.get(tc.get("quality_status", ""))
                if q_color:
                    color_requests.append({"range": f"N{4 + i}", "format": {"backgroundColor": q_color}})
            if color_requests:
                ws.batch_format(color_requests)

            # 기존 드롭다운 초기화 후 테스트 상태(I열)/자동화 가능여부(M열)/Quality 판정(N열) 드롭다운 재설정
            sh.batch_update({"requests": [
                {
                    "setDataValidation": {
                        "range": {
                            "sheetId": ws.id,
                            "startRowIndex": 3,
                            "endRowIndex": end_row,
                            "startColumnIndex": 0,
                            "endColumnIndex": 21,
                        },
                    }
                },
                {
                    "setDataValidation": {
                        "range": {
                            "sheetId": ws.id,
                            "startRowIndex": 3,
                            "endRowIndex": end_row,
                            "startColumnIndex": 8,
                            "endColumnIndex": 9,
                        },
                        "rule": {
                            "condition": {
                                "type": "ONE_OF_LIST",
                                "values": [
                                    {"userEnteredValue": "P"},
                                    {"userEnteredValue": "F"},
                                    {"userEnteredValue": "B"},
                                    {"userEnteredValue": "N/A"},
                                ],
                            },
                            "showCustomUi": True,
                            "strict": False,
                        },
                    }
                },
                {
                    "setDataValidation": {
                        "range": {
                            "sheetId": ws.id,
                            "startRowIndex": 3,
                            "endRowIndex": end_row,
                            "startColumnIndex": 12,
                            "endColumnIndex": 13,
                        },
                        "rule": {
                            "condition": {
                                "type": "ONE_OF_LIST",
                                "values": [
                                    {"userEnteredValue": "가능"},
                                    {"userEnteredValue": "불가능"},
                                ],
                            },
                            "showCustomUi": True,
                            "strict": False,
                        },
                    }
                },
                {
                    "setDataValidation": {
                        "range": {
                            "sheetId": ws.id,
                            "startRowIndex": 3,
                            "endRowIndex": end_row,
                            "startColumnIndex": 13,
                            "endColumnIndex": 14,
                        },
                        "rule": {
                            "condition": {
                                "type": "ONE_OF_LIST",
                                "values": [
                                    {"userEnteredValue": "PASS"},
                                    {"userEnteredValue": "REVIEW"},
                                    {"userEnteredValue": "REJECT"},
                                ],
                            },
                            "showCustomUi": True,
                            "strict": False,
                        },
                    }
                },
            ]})

            # 열 너비 설정
            requests_body = [{"updateDimensionProperties": {
                "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": i, "endIndex": i + 1},
                "properties": {"pixelSize": px},
                "fields": "pixelSize",
            }} for i, px in enumerate(col_widths)]
            sh.batch_update({"requests": requests_body})

            # 자동화가능여부(M, index 12) + 요구사항/실행성/설계기법 메타데이터(P~U, index 15~20)는
            # 수동 테스트 수행 시 방해되지 않도록 기본 숨김 처리 (Quality 판정/사유는 N/O로 노출 유지)
            hidden_ranges = [(12, 13), (15, 21)]
            sh.batch_update({"requests": [
                {
                    "updateDimensionProperties": {
                        "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": start, "endIndex": end},
                        "properties": {"hiddenByUser": True},
                        "fields": "hiddenByUser",
                    }
                } for start, end in hidden_ranges
            ]})

        total_tc += len(item["test_cases"])
        print(f"  '{sheet_title}' — TC {len(item['test_cases'])}개 저장")

    sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}"
    print(f"\n  총 {total_tc}개 TC 저장 완료")
    print(f"  {sheet_url}")
    return sheet_url


def process_keys(jira: JIRA, groq_client: Groq, issue_keys: list, context: str = "") -> list:
    """티켓 키 목록을 순서대로 처리하여 결과를 반환합니다."""
    results = []
    total = len(issue_keys)

    for idx, key in enumerate(issue_keys, start=1):
        print(f"\n[{idx}/{total}] {key} 처리 중...")

        try:
            issue = fetch_issue(jira, key)
        except Exception as e:
            print(f"  [건너뜀] 티켓 조회 실패: {e}")
            continue

        print(f"  제목: {issue['summary']} | 상태: {issue['status']}")
        print(f"  요구사항 추론 중...")
        try:
            augmented_spec = augment_ticket_spec(groq_client, issue, context)
        except DailyTokenLimitError as e:
            print(f"  [일일 한도 초과] {e}")
            print(f"  지금까지 처리된 {len(results)}개 티켓 결과로 저장합니다.")
            break
        print(f"  --- 요구사항 분석 ---\n{augmented_spec}\n  ---")
        print(f"  TC 생성 중...")
        try:
            tc_list = generate_test_cases(groq_client, issue, augmented_spec, context)
        except DailyTokenLimitError as e:
            print(f"  [일일 한도 초과] {e}")
            print(f"  지금까지 처리된 {len(results)}개 티켓 결과로 저장합니다.")
            break
        tc_list = filter_tc_list(tc_list)
        tc_list = dedupe_tc_list(tc_list)
        print(f"  생성된 TC: {len(tc_list)}개")
        for tc in tc_list:
            print(f"    [{tc.get('tc_id')}] [{tc.get('대분류', '-')}] [{tc.get('테스트유형', '-')}] [{tc.get('우선순위', '-')}] {tc.get('테스트시나리오', '')}")

        results.append({
            "key": issue["key"],
            "summary": issue["summary"],
            "status": issue["status"],
            "augmented_spec": augmented_spec,
            "test_cases": tc_list,
        })

        if idx < total:
            import time
            print(f"  다음 티켓까지 30초 대기...")
            time.sleep(30)

    return results


# ── 메인 ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("input", nargs="?", default=None)
    parser.add_argument("--context", default="", help="서비스 컨텍스트 이름 (예: kream, kurly)")
    parser.add_argument("--template", action="store_true")
    args, _ = parser.parse_known_args()

    if not args.input and not args.template:
        print("사용법:")
        print("  python src/generate_tc_from_url.py MKQA-1")
        print("  python src/generate_tc_from_url.py MKQA-1 --context kream")
        print("  python src/generate_tc_from_url.py tickets.xlsx --context kurly")
        print("  python src/generate_tc_from_url.py --template")
        sys.exit(1)

    # 컨텍스트 로드
    context = load_context(args.context)
    if context:
        print(f"  컨텍스트 로드됨: contexts/{args.context}.md")

    input_str = args.input

    # ── 템플릿 생성 모드
    if args.template:
        path = create_template("tickets_template.xlsx")
        print(f"템플릿 생성 완료: {path}")
        print("A열에 티켓 URL 또는 이슈 키를 입력한 후 실행하세요.")
        return

    # ── 구글 스프레드시트 모드
    if "docs.google.com/spreadsheets" in input_str or re.fullmatch(r"[a-zA-Z0-9_-]{44}", input_str.strip()):
        try:
            sheet_id = extract_sheet_id(input_str)
        except ValueError as e:
            print(f"[오류] {e}")
            sys.exit(1)
        gid = extract_gid(input_str)

        print(f"\n=== 구글 스프레드시트에서 티켓 목록 읽는 중... ===")
        if gid:
            print(f"  대상 탭: gid={gid}")
        raw_keys = read_keys_from_sheets(sheet_id, gid)
        if not raw_keys:
            print("[오류] 스프레드시트 A열(2행~)에 티켓 URL/키가 없습니다.")
            sys.exit(1)

        issue_keys = []
        for raw in raw_keys:
            try:
                issue_keys.append(extract_issue_key(raw))
            except ValueError:
                print(f"  [건너뜀] 유효하지 않은 값: {raw}")

        if not issue_keys:
            print("[오류] 유효한 이슈 키가 없습니다.")
            sys.exit(1)

        print(f"  읽어온 티켓: {len(issue_keys)}개 → {', '.join(issue_keys)}")
        print(f"\n=== TC 생성 시작 ===")
        label = f"sheets_{sheet_id[:8]}"
        use_sheets = True

    # ── 엑셀 파일 모드
    elif input_str.endswith(".xlsx") or input_str.endswith(".xls"):
        if not os.path.exists(input_str):
            print(f"[오류] 파일을 찾을 수 없습니다: {input_str}")
            sys.exit(1)

        raw_keys = read_keys_from_excel(input_str)
        if not raw_keys:
            print("[오류] 엑셀 A열에 티켓 URL/키가 없습니다.")
            sys.exit(1)

        issue_keys = []
        for raw in raw_keys:
            try:
                issue_keys.append(extract_issue_key(raw))
            except ValueError:
                print(f"  [건너뜀] 유효하지 않은 값: {raw}")

        if not issue_keys:
            print("[오류] 유효한 이슈 키가 없습니다.")
            sys.exit(1)

        print(f"\n=== 엑셀 일괄 TC 생성 시작: {len(issue_keys)}개 티켓 ===")
        label = f"batch_{os.path.splitext(os.path.basename(input_str))[0]}"
        sheet_id = None
        use_sheets = False

    # ── 단일 티켓 모드
    else:
        try:
            issue_key = extract_issue_key(input_str)
        except ValueError as e:
            print(f"[오류] {e}")
            sys.exit(1)
        issue_keys = [issue_key]
        print(f"\n=== 티켓 TC 자동 생성 시작: {issue_key} ===")
        label = issue_key.replace("/", "_")
        sheet_id = None
        use_sheets = False

    # 클라이언트 초기화
    jira = JIRA(
        server=os.getenv("JIRA_URL"),
        basic_auth=(os.getenv("JIRA_EMAIL"), os.getenv("JIRA_API_TOKEN")),
    )
    groq_client = Groq(api_key=os.getenv("GROQ_API_KEY"))

    # TC 생성
    results = process_keys(jira, groq_client, issue_keys, context)

    if not results:
        print("\n[오류] 생성된 TC가 없습니다.")
        sys.exit(1)

    # 결과 저장
    print(f"\n=== 결과 저장 중... ===")
    os.makedirs("reports", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    json_path = f"reports/tc_{label}_{timestamp}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"  JSON 저장 완료: {json_path}")

    xlsx_path = f"reports/tc_{label}_{timestamp}.xlsx"
    save_excel(results, xlsx_path)
    print(f"  엑셀 저장 완료: {xlsx_path}")

    # 구글 시트 모드일 경우 시트에도 저장
    if use_sheets:
        save_to_sheets(results, sheet_id)

    total_tc = sum(len(r["test_cases"]) for r in results)
    print(f"\n=== 완료: {len(results)}개 티켓 / {total_tc}개 TC 생성 ===")


if __name__ == "__main__":
    main()
